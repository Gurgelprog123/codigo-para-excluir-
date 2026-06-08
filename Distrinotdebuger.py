from DijurLib.api.base import post_api_navegador, get_api_navegador, put_api_navegador

# Biblioteca Dijur
from DijurLib.portal.login import login_manual
from DijurLib.api.helpers import get_logged_user

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.edge.service import Service

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

from datetime import datetime

import pandas as pd
import os
import sys
import time
import re
import base64
import unicodedata
import json
import traceback


# ==========================================================
# CONFIGURACOES GERAIS
# ==========================================================

TEMPO = 1.5

# Mantive como caracteres porque era assim que seu codigo validava.
# Se quiser usar linhas reais, troque len(texto) por len(texto.splitlines()).
LIMITE_CARACTERES_PUBLICACAO = 3000

TEXTO_FIXO_DISTRIBUICAO_ADVOGADO = "NPJ distribuído para o advogado nesta data."

URL_CONSULTAR_EVENTOS = "https://juridico.intranet.bb.com.br/paj/resources/app/v1/agenda/consultar/evento/compromisso/responsavel/"
URL_ALTERAR_ESTADO_EVENTO = "https://juridico.intranet.bb.com.br/paj/resources/app/v1/agenda/evento/estado"
URL_REGISTRAR_OBSERVACAO = "https://juridico.intranet.bb.com.br/paj/resources/app/v1/agenda/evento/observacao"
URL_PUBLICACAO = "https://juridico.intranet.bb.com.br/paj/resources/app/v1/agenda/publicacao"


# ==========================================================
# DIRETORIOS / LOG
# ==========================================================

def obter_diretorio_execucao():
    """
    Retorna o diretorio correto para salvar relatorio e log.
    Quando virar .exe, salva ao lado do executavel.
    Quando rodar como .py, salva na pasta do script.
    """

    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)

    return os.path.dirname(os.path.abspath(__file__))


DIRETORIO_EXECUCAO = obter_diretorio_execucao()

NOME_LOG = f"log_execucao_baixas_{datetime.now().strftime('%d-%m-%Y_%H-%M-%S')}.txt"
CAMINHO_LOG = os.path.join(DIRETORIO_EXECUCAO, NOME_LOG)


def logar(mensagem):
    """
    Registra mensagem no console e em arquivo de log.
    """

    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    linha = f"[{data_hora}] {mensagem}"

    print(linha)

    try:
        with open(CAMINHO_LOG, "a", encoding="utf-8") as arquivo:
            arquivo.write(linha + "\n")
    except Exception:
        pass


def mostrar_mensagem_final(titulo, mensagem, tipo="info"):
    """
    Mostra uma mensagem final simples.
    Nao e debugger: apenas aviso de conclusao/erro para uso em .exe.
    """

    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        if tipo == "erro":
            messagebox.showerror(titulo, mensagem, parent=root)
        elif tipo == "aviso":
            messagebox.showwarning(titulo, mensagem, parent=root)
        else:
            messagebox.showinfo(titulo, mensagem, parent=root)

        root.destroy()

    except Exception:
        pass


# ==========================================================
# NAVEGADOR
# ==========================================================

def iniciar_navegador(headless: bool = False, wait_time: int = 30, suppress_browser_ui: bool = False):
    """
    Inicializa o navegador Edge com as opcoes especificadas.
    """

    options = webdriver.EdgeOptions()

    if headless:
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    prefs = {
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True
    }

    if suppress_browser_ui:
        options.add_argument("--disable-notifications")
        options.add_argument("--no-download-notification")

        prefs.update({
            "browser.show_hub_popup_on_download_start": False,
            "user_experience_metrics.personalization_data_consent_enabled": True
        })

    options.add_experimental_option("prefs", prefs)

    try:
        if os.path.exists("C:\\EdgeDriver\\msedgedriver.exe"):
            service = Service(executable_path="C:\\EdgeDriver\\msedgedriver.exe")
            driver = webdriver.Edge(service=service, options=options)
        else:
            driver = webdriver.Edge(options=options)

    except Exception as e:
        msg = str(e)

        match = re.search(
            r"only supports Microsoft Edge version (\d+).*?Current browser version is (\d+\.\d+\.\d+\.\d+)",
            msg,
            re.DOTALL
        )

        caminho_erro = os.path.join(DIRETORIO_EXECUCAO, "navegador_erro_log.txt")

        with open(caminho_erro, "w", encoding="utf-8") as f:
            if match:
                versao_suportada = match.group(1)
                versao_atual = match.group(2)

                f.write(
                    "Incompatibilidade de versao do Edge WebDriver.\n"
                    f"Driver suporta apenas a versao: {versao_suportada}\n"
                    f"Sua versao atual do Edge e: {versao_atual}\n"
                    "Atualize o msedgedriver.exe conforme a versao atual do navegador.\n"
                )

                try:
                    pdf_b64 = b"""
                    JJABDUIWDVBAPWVWI
                    """
                    nome_arquivo = os.path.join(DIRETORIO_EXECUCAO, "Manual Baixar Driver.pdf")

                    with open(nome_arquivo, "wb") as arquivo_pdf:
                        arquivo_pdf.write(base64.b64decode(pdf_b64))

                except Exception:
                    f.write("\nNao foi possivel gerar o PDF de orientacao automaticamente.\n")

            else:
                f.write(f"Ocorreu um erro ao abrir o navegador:\n{msg}")

        raise Exception("Erro de navegador. Verifique o arquivo navegador_erro_log.txt.")

    driver.maximize_window()

    wait = WebDriverWait(driver, wait_time)

    return driver, wait


# ==========================================================
# FUNCOES AUXILIARES
# ==========================================================

def serializar_objeto(objeto, limite=5000):
    """
    Transforma resposta/payload em texto legivel para relatorio.
    """

    try:
        texto = json.dumps(objeto, ensure_ascii=False, indent=2, default=str)
    except Exception:
        texto = str(objeto)

    if texto is None:
        texto = ""

    if len(texto) > limite:
        texto = texto[:limite] + "\n\n... TEXTO CORTADO PARA VISUALIZACAO ..."

    return texto


def normalizar_texto(texto):
    """
    Normaliza texto para comparacoes mais seguras.
    """

    if texto is None:
        return ""

    texto = str(texto).lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(char for char in texto if unicodedata.category(char) != "Mn")
    texto = re.sub(r"\s+", " ", texto).strip()

    return texto


def eh_texto_fixo_distribuicao_advogado(texto_descricao_compromisso):
    """
    Verifica a primeira forma de baixa:
    texto fixo de NPJ distribuido para advogado.
    """

    texto_recebido = normalizar_texto(texto_descricao_compromisso)
    texto_esperado = normalizar_texto(TEXTO_FIXO_DISTRIBUICAO_ADVOGADO)

    return texto_recebido == texto_esperado


def extrair_turma_distribuicao_tst(conteudo_publicacao):
    """
    Verifica a segunda forma de baixa:
    publicacao curta do TST informando distribuicao para turma.
    """

    texto = normalizar_texto(conteudo_publicacao)

    padrao = (
        r"\bdistribuid[oa]\s+para\s+"
        r"(?:a\s+)?"
        r"(\d{1,2})"
        r"\s*(?:a|o|ª|º)?\s*"
        r"turma\b"
    )

    match = re.search(padrao, texto, flags=re.IGNORECASE)

    if not match:
        return None

    numero_turma = int(match.group(1))

    if numero_turma <= 0:
        return None

    return f"{numero_turma:02d} TURMA"


def resposta_sucesso(resposta):
    """
    Padroniza a validacao de sucesso das respostas da API.
    """

    if not resposta:
        return False

    if resposta.get("type") == "SUCCESS":
        return True

    if resposta.get("status") == "OK":
        return True

    return False


def resumo_publicacao(texto, limite=500):
    """
    Gera resumo curto da publicacao para o relatorio.
    """

    if not texto:
        return ""

    texto = re.sub(r"\s+", " ", str(texto)).strip()

    if len(texto) > limite:
        return texto[:limite] + "..."

    return texto


def criar_registro_relatorio(evento, indice_evento):
    """
    Cria registro base para auditoria do evento.
    """

    return {
        "Data Hora Execucao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "Indice Evento": indice_evento,
        "Codigo Evento": evento.get("codigoEventoJuridico"),
        "NPJ": evento.get("numeroProcessoFormatado"),
        "Texto Descricao Compromisso": evento.get("textoDescricaoCompromisso"),
        "Status Final": "NAO PROCESSADO",
        "Regra Identificada": "",
        "Acao Executada": "",
        "Motivo": "",
        "Tramitacao Identificada": "",
        "Numero Processo": "",
        "Numero Sequencial Publicacao": "",
        "Qtd Caracteres Publicacao": "",
        "Qtd Linhas Publicacao": "",
        "Resumo Publicacao": "",
        "Resposta PUT": "",
        "Resposta Observacao": "",
        "Erro": ""
    }


# ==========================================================
# POST PLAIN
# ==========================================================

def post_api_navegador_plain(driver, api_url: str, payload: str, max_attempts: int = 3):
    """
    Faz POST enviando payload como texto puro.
    Usado na consulta da publicacao.
    """

    def parse_service_response(texto: str) -> dict:
        padrao = r"status=(\w+).*?messages?=\[([^\]]*)\].*?data=(.*)"
        match = re.search(padrao, texto, flags=re.DOTALL)

        if match:
            status, message, data = match.groups()

            return {
                "status": status,
                "message": message,
                "data": data
            }

        return {"raw": texto}

    attempts = 0

    while attempts < max_attempts:
        try:
            logar(f"Iniciando requisicao POST plain. Tentativa {attempts + 1}.")

            response_data = driver.execute_async_script(
                """
                const api_url = arguments[0];
                const payload = arguments[1];
                const callback = arguments[arguments.length - 1];

                fetch(api_url, {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: payload,
                    credentials: 'same-origin'
                })
                .then(response => {
                    if (!response.ok) {
                        throw new Error('Network response was not ok: ' + response.statusText);
                    }
                    return response.text();
                })
                .then(text => {
                    try {
                        const parsed = JSON.parse(text);
                        callback(parsed);
                    } catch (err) {
                        callback({ rawText: text });
                    }
                })
                .catch(error => callback({ error: error.toString() }));
                """,
                api_url,
                payload
            )

            if "error" in response_data:
                logar(f"Erro na requisicao POST plain: {response_data['error']}")
                return None

            if "rawText" in response_data:
                parsed = parse_service_response(response_data["rawText"])
                logar("Requisicao POST plain bem-sucedida em formato nao JSON.")
                return parsed

            logar("Requisicao POST plain bem-sucedida em JSON.")
            return response_data

        except Exception as e:
            logar("Erro ao fazer requisicao POST plain.")
            logar(str(e))
            logar(traceback.format_exc())

        attempts += 1

        if attempts < max_attempts:
            logar("Tentando novamente em 1 segundo...")
            time.sleep(1)

    logar("Numero maximo de tentativas alcancado no POST plain.")
    return None


# ==========================================================
# API / FLUXO DE NEGOCIO
# ==========================================================

def buscar_eventos_pendentes(driver, matricula):
    """
    Busca eventos pendentes da agenda.
    """

    todos_eventos = []
    posicao = 1

    while True:
        payload = {
            "posicao": posicao,
            "responsavel": matricula,
            "situacao": 5
        }

        logar(f"Buscando eventos pendentes. Posicao: {posicao}")

        resposta = post_api_navegador(driver, URL_CONSULTAR_EVENTOS, payload)

        if resposta is None or resposta.get("status") != "OK":
            logar("Status na busca de publicacoes veio None ou diferente de OK.")
            break

        lista_eventos = resposta.get("data", {}).get("listaEvento", [])

        if not lista_eventos:
            logar("Nenhum evento retornado nesta posicao.")
            break

        todos_eventos.extend(lista_eventos)

        logar(f"Eventos encontrados nesta posicao: {len(lista_eventos)}")

        if len(lista_eventos) < 50:
            break

        posicao += 50
        time.sleep(TEMPO)

    return todos_eventos


def obter_numero_processo(driver, num_processo_formatado):
    """
    Obtem numeroProcesso a partir do NPJ formatado.
    """

    numero_processo_sem_barra = num_processo_formatado.replace("/", "")
    partes = numero_processo_sem_barra.split("-")

    if len(partes) != 2:
        logar(f"NPJ em formato inesperado: {num_processo_formatado}")
        return None, None

    npj_base = partes[0]

    try:
        npj_sufixo = int(partes[1])

    except ValueError:
        logar(f"Variacao do NPJ invalida: {num_processo_formatado}")
        return None, None

    url_numeroprocesso = (
        "https://juridico.intranet.bb.com.br/paj/resources/app/v1/"
        f"processo/consulta/{npj_base}/{npj_sufixo}/0"
    )

    logar(f"Buscando numeroProcesso do NPJ {num_processo_formatado}.")

    resposta_numeroprocesso = get_api_navegador(driver, url_numeroprocesso)

    if resposta_numeroprocesso is None or resposta_numeroprocesso.get("status") != "OK":
        logar("Status da chamada para pegar numeroProcesso diferente de OK.")
        return None, resposta_numeroprocesso

    lista_ocorrencia = resposta_numeroprocesso.get("data", {}).get("listaOcorrencia", [])

    if not lista_ocorrencia:
        logar(f"Nenhuma ocorrencia encontrada para o processo: {num_processo_formatado}")
        return None, resposta_numeroprocesso

    numero_processo = lista_ocorrencia[0].get("numeroProcesso")

    if not numero_processo:
        logar(f"numeroProcesso nao encontrado para: {num_processo_formatado}")
        return None, resposta_numeroprocesso

    return numero_processo, resposta_numeroprocesso


def obter_numero_sequencial_publicacao(driver, cod_evento, matricula, npj):
    """
    Obtem numeroSequencialPublicacaoJudicial pelo codigo do evento.
    """

    url_detalhes = (
        "https://juridico.intranet.bb.com.br/paj/resources/app/v1/"
        f"agenda/evento/{cod_evento}/{matricula}"
    )

    logar(f"Buscando detalhes do evento {cod_evento} / NPJ {npj}.")

    resposta_detalhes = get_api_navegador(driver, url_detalhes)

    if resposta_detalhes is None or resposta_detalhes.get("status") != "OK":
        logar("Status da chamada para pegar numeroSequencialPublicacaoJudicial diferente de OK.")
        return None, resposta_detalhes

    numero_sequencial = resposta_detalhes.get("data", {}).get("numeroSequencialPublicacaoJudicial")

    if not numero_sequencial:
        logar(f"numeroSequencialPublicacaoJudicial nao encontrado para evento: {cod_evento}")
        return None, resposta_detalhes

    return numero_sequencial, resposta_detalhes


def obter_conteudo_publicacao(driver, cod_evento, numero_sequencial_publicacao, npj):
    """
    Busca o conteudo da publicacao.
    """

    payload_publicacao = f"{numero_sequencial_publicacao};{cod_evento}"

    logar(f"Buscando conteudo da publicacao do NPJ {npj}.")

    resposta_publicacao = post_api_navegador_plain(
        driver=driver,
        api_url=URL_PUBLICACAO,
        payload=payload_publicacao
    )

    if resposta_publicacao is None:
        logar("Resposta da chamada de publicacao veio None.")
        return None, resposta_publicacao

    status_publicacao = resposta_publicacao.get("status")
    type_publicacao = resposta_publicacao.get("type")

    if status_publicacao != "OK" and type_publicacao != "SUCCESS":
        logar("Status da chamada para pegar conteudo diferente de OK/SUCCESS.")
        logar(serializar_objeto(resposta_publicacao, limite=1500))
        return None, resposta_publicacao

    dados_publicacao = resposta_publicacao.get("data")

    if isinstance(dados_publicacao, dict):
        conteudo_publicacao = dados_publicacao.get("textoPublicacao")

    elif isinstance(dados_publicacao, str):
        conteudo_publicacao = dados_publicacao

    else:
        logar("Resposta da publicacao veio em formato inesperado.")
        logar(serializar_objeto(resposta_publicacao, limite=1500))
        return None, resposta_publicacao

    if not conteudo_publicacao:
        logar("Publicacao sem texto.")
        return None, resposta_publicacao

    return conteudo_publicacao, resposta_publicacao


def concluir_evento(driver, matricula, cod_evento, motivo, npj, regra):
    """
    Realiza o PUT que conclui o evento.
    Esse e o ponto onde a baixa acontece.
    """

    put_payload = {
        "chave": matricula,
        "codigoEventoJuridico": cod_evento,
        "codigoTipoEstadoCompromisso": 3,
        "inclurAndamentoAutomatico": True,
        "justificativaConclusao": motivo,
    }

    logar(f"Realizando PUT de baixa. Evento {cod_evento} / NPJ {npj} / Regra {regra}.")

    resposta_put = put_api_navegador(
        driver,
        URL_ALTERAR_ESTADO_EVENTO,
        put_payload
    )

    logar(f"Resposta PUT: {serializar_objeto(resposta_put, limite=1000)}")

    return resposta_put


def registrar_observacao_conclusao(driver, cod_evento, texto_observacao, npj):
    """
    Registra observacao depois que a baixa foi feita com sucesso.
    """

    payload_post_descricao = {
        "codigoEventoJuridico": cod_evento,
        "textoObservacaoCompromisso": texto_observacao
    }

    logar(f"Registrando observacao de conclusao. Evento {cod_evento} / NPJ {npj}.")

    resposta_observacao = post_api_navegador(
        driver,
        URL_REGISTRAR_OBSERVACAO,
        payload_post_descricao
    )

    logar(f"Resposta observacao: {serializar_objeto(resposta_observacao, limite=1000)}")

    return resposta_observacao


# ==========================================================
# RELATORIO
# ==========================================================

def salvar_relatorio_excel(relatorio_dados):
    """
    Gera relatorio Excel com auditoria completa da execucao.
    """

    if not relatorio_dados:
        logar("Nenhum dado coletado para gerar o relatorio.")
        return None

    df_relatorio = pd.DataFrame(relatorio_dados)

    data_hoje = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    nome_arquivo = f"Relatorio auditoria baixas distribuicao {data_hoje}.xlsx"
    caminho_arquivo = os.path.join(DIRETORIO_EXECUCAO, nome_arquivo)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Auditoria Baixas"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, column_name in enumerate(df_relatorio.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment

        for row_idx, row in enumerate(dataframe_to_rows(df_relatorio, index=False, header=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = alignment

                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    tamanho = len(str(cell.value))

                    if tamanho > max_length:
                        max_length = tamanho

                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 35

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(caminho_arquivo)

        logar(f"Relatorio de auditoria salvo com sucesso em: {caminho_arquivo}")

        return caminho_arquivo

    except Exception as e:
        logar("Erro ao salvar o relatorio Excel.")
        logar(str(e))
        logar(traceback.format_exc())

        return None


# ==========================================================
# FUNCAO PRINCIPAL
# ==========================================================

def organizar_tramitacao_agenda_gerente(driver, matricula):
    """
    Organiza a agenda e realiza baixa de publicacoes pendentes de prazo
    em duas situacoes:

    1. Evento com texto fixo:
       NPJ distribuido para o advogado nesta data.

    2. Publicacao curta do TST:
       texto com ate 3000 caracteres e regex de distribuicao para turma.
    """

    dados_processos = []
    relatorio_dados = []
    eventos_baixados = set()

    try:
        todos_eventos = buscar_eventos_pendentes(driver, matricula)

        if not todos_eventos:
            logar("Nenhum evento encontrado ou erro na resposta.")
            return dados_processos, relatorio_dados

        logar("Agenda listada.")
        logar(f"Quantidade de eventos encontrados: {len(todos_eventos)}")

        for indice_evento, evento in enumerate(todos_eventos, 1):
            registro = criar_registro_relatorio(evento, indice_evento)

            try:
                cod_evento = evento.get("codigoEventoJuridico")
                num_processo_formatado = evento.get("numeroProcessoFormatado")
                texto_descricao_compromisso = evento.get("textoDescricaoCompromisso")

                logar("=" * 80)
                logar(f"Processando evento {indice_evento}/{len(todos_eventos)}")
                logar(f"Codigo Evento: {cod_evento}")
                logar(f"NPJ: {num_processo_formatado}")

                if not cod_evento:
                    registro["Status Final"] = "NAO BAIXADO - SEM CODIGO EVENTO"
                    registro["Acao Executada"] = "Nenhuma"
                    registro["Motivo"] = "Evento sem codigoEventoJuridico."
                    continue

                if cod_evento in eventos_baixados:
                    registro["Status Final"] = "NAO BAIXADO - EVENTO JA BAIXADO NA EXECUCAO"
                    registro["Acao Executada"] = "Nenhuma"
                    registro["Motivo"] = "Evento ja foi baixado anteriormente nesta mesma execucao."
                    continue

                # ==================================================
                # REGRA 1: BAIXA PELO TEXTO FIXO DO COMPROMISSO
                # ==================================================

                if eh_texto_fixo_distribuicao_advogado(texto_descricao_compromisso):
                    registro["Regra Identificada"] = "Distribuicao interna - texto fixo"
                    registro["Tramitacao Identificada"] = "DISTRIBUICAO PARA ADVOGADO"

                    logar("Evento identificado pela regra 1: texto fixo de distribuicao para advogado.")

                    resposta_put = concluir_evento(
                        driver=driver,
                        matricula=matricula,
                        cod_evento=cod_evento,
                        motivo="Distribuição.",
                        npj=num_processo_formatado,
                        regra="Distribuicao interna - texto fixo"
                    )

                    registro["Resposta PUT"] = serializar_objeto(resposta_put, limite=1500)

                    if not resposta_sucesso(resposta_put):
                        registro["Status Final"] = "ERRO - FALHA NO PUT"
                        registro["Acao Executada"] = "Tentou realizar baixa"
                        registro["Motivo"] = "PUT retornou erro ou resposta diferente de sucesso."
                        logar("Falha ao concluir evento pela regra 1. Execucao interrompida por seguranca.")
                        break

                    eventos_baixados.add(cod_evento)

                    time.sleep(TEMPO)

                    resposta_observacao = registrar_observacao_conclusao(
                        driver=driver,
                        cod_evento=cod_evento,
                        npj=num_processo_formatado,
                        texto_observacao=(
                            "Evento concluido. Motivo: Distribuição. "
                            "Regra: texto fixo NPJ distribuído para o advogado."
                        )
                    )

                    registro["Resposta Observacao"] = serializar_objeto(resposta_observacao, limite=1500)

                    if not resposta_sucesso(resposta_observacao):
                        registro["Status Final"] = "ERRO - FALHA AO REGISTRAR OBSERVACAO"
                        registro["Acao Executada"] = "Baixou evento, mas falhou ao registrar observacao"
                        registro["Motivo"] = "A baixa foi feita, mas a observacao retornou erro."
                        logar("Falha ao registrar observacao pela regra 1. Execucao interrompida por seguranca.")
                        break

                    registro["Status Final"] = "BAIXADO - DISTRIBUICAO INTERNA"
                    registro["Acao Executada"] = "Baixa realizada por PUT e observacao registrada"
                    registro["Motivo"] = "Texto fixo de distribuicao interna identificado."

                    logar("Evento baixado pela regra 1 com sucesso.")

                    time.sleep(TEMPO)

                    continue

                # ==================================================
                # REGRA 2: BAIXA POR PUBLICACAO CURTA DO TST
                # ==================================================

                if not num_processo_formatado:
                    registro["Status Final"] = "NAO BAIXADO - SEM NPJ"
                    registro["Acao Executada"] = "Nenhuma"
                    registro["Motivo"] = "Evento sem numeroProcessoFormatado."
                    continue

                numero_processo, resposta_numero_processo = obter_numero_processo(
                    driver=driver,
                    num_processo_formatado=num_processo_formatado
                )

                if not numero_processo:
                    registro["Status Final"] = "NAO BAIXADO - PROCESSO NAO LOCALIZADO"
                    registro["Acao Executada"] = "Consultou processo"
                    registro["Motivo"] = "Nao foi possivel obter numeroProcesso."
                    continue

                registro["Numero Processo"] = numero_processo

                time.sleep(TEMPO)

                numero_sequencial_publicacao, resposta_detalhes = obter_numero_sequencial_publicacao(
                    driver=driver,
                    cod_evento=cod_evento,
                    matricula=matricula,
                    npj=num_processo_formatado
                )

                if not numero_sequencial_publicacao:
                    registro["Status Final"] = "NAO BAIXADO - SEM PUBLICACAO JUDICIAL"
                    registro["Acao Executada"] = "Consultou detalhes do evento"
                    registro["Motivo"] = "Nao encontrou numeroSequencialPublicacaoJudicial."
                    continue

                registro["Numero Sequencial Publicacao"] = numero_sequencial_publicacao

                time.sleep(TEMPO)

                conteudo_publicacao, resposta_publicacao = obter_conteudo_publicacao(
                    driver=driver,
                    cod_evento=cod_evento,
                    numero_sequencial_publicacao=numero_sequencial_publicacao,
                    npj=num_processo_formatado
                )

                if not conteudo_publicacao:
                    registro["Status Final"] = "NAO BAIXADO - SEM TEXTO DA PUBLICACAO"
                    registro["Acao Executada"] = "Consultou publicacao"
                    registro["Motivo"] = "Nao foi possivel obter texto da publicacao."
                    continue

                quantidade_caracteres = len(conteudo_publicacao)
                quantidade_linhas = len(str(conteudo_publicacao).splitlines())

                registro["Qtd Caracteres Publicacao"] = quantidade_caracteres
                registro["Qtd Linhas Publicacao"] = quantidade_linhas
                registro["Resumo Publicacao"] = resumo_publicacao(conteudo_publicacao, limite=500)

                if quantidade_caracteres > LIMITE_CARACTERES_PUBLICACAO:
                    registro["Status Final"] = "NAO BAIXADO - PUBLICACAO EXTENSA"
                    registro["Acao Executada"] = "Consultou publicacao, mas nao baixou"
                    registro["Motivo"] = (
                        f"Publicacao com {quantidade_caracteres} caracteres. "
                        f"Limite configurado: {LIMITE_CARACTERES_PUBLICACAO}."
                    )

                    logar(
                        "Publicacao ignorada por tamanho. "
                        f"Caracteres: {quantidade_caracteres}. "
                        f"Limite: {LIMITE_CARACTERES_PUBLICACAO}."
                    )

                    continue

                tramitacao_distribuida = extrair_turma_distribuicao_tst(conteudo_publicacao)

                if not tramitacao_distribuida:
                    registro["Status Final"] = "NAO BAIXADO - NAO ERA DISTRIBUICAO"
                    registro["Regra Identificada"] = "Nenhuma"
                    registro["Acao Executada"] = "Consultou publicacao, mas nao baixou"
                    registro["Motivo"] = "Publicacao curta, mas sem regex de distribuicao para turma."

                    logar("Publicacao curta, mas sem regex de distribuicao para turma.")

                    continue

                registro["Regra Identificada"] = "Distribuicao TST - publicacao curta"
                registro["Tramitacao Identificada"] = tramitacao_distribuida

                logar(f"Evento identificado pela regra 2: publicacao TST distribuida para {tramitacao_distribuida}.")

                resposta_put = concluir_evento(
                    driver=driver,
                    matricula=matricula,
                    cod_evento=cod_evento,
                    motivo="Distribuição.",
                    npj=num_processo_formatado,
                    regra="Distribuicao TST - publicacao curta"
                )

                registro["Resposta PUT"] = serializar_objeto(resposta_put, limite=1500)

                if not resposta_sucesso(resposta_put):
                    registro["Status Final"] = "ERRO - FALHA NO PUT"
                    registro["Acao Executada"] = "Tentou realizar baixa"
                    registro["Motivo"] = "PUT retornou erro ou resposta diferente de sucesso."
                    logar("Falha ao concluir evento pela regra 2. Execucao interrompida por seguranca.")
                    break

                eventos_baixados.add(cod_evento)

                time.sleep(TEMPO)

                resposta_observacao = registrar_observacao_conclusao(
                    driver=driver,
                    cod_evento=cod_evento,
                    npj=num_processo_formatado,
                    texto_observacao=(
                        "Evento concluido. Motivo: Distribuição. "
                        f"Regra: publicação TST curta. Tramitação identificada: {tramitacao_distribuida}."
                    )
                )

                registro["Resposta Observacao"] = serializar_objeto(resposta_observacao, limite=1500)

                if not resposta_sucesso(resposta_observacao):
                    registro["Status Final"] = "ERRO - FALHA AO REGISTRAR OBSERVACAO"
                    registro["Acao Executada"] = "Baixou evento, mas falhou ao registrar observacao"
                    registro["Motivo"] = "A baixa foi feita, mas a observacao retornou erro."
                    logar("Falha ao registrar observacao pela regra 2. Execucao interrompida por seguranca.")
                    break

                registro["Status Final"] = "BAIXADO - DISTRIBUICAO TST"
                registro["Acao Executada"] = "Baixa realizada por PUT e observacao registrada"
                registro["Motivo"] = "Publicacao curta do TST com regex de distribuicao para turma."

                logar("Evento baixado pela regra 2 com sucesso.")

                time.sleep(TEMPO)

            except Exception as e:
                registro["Status Final"] = "ERRO"
                registro["Acao Executada"] = "Erro durante processamento do evento"
                registro["Erro"] = str(e)
                registro["Motivo"] = "Ocorreu erro inesperado no processamento deste evento."

                logar(f"Erro ao processar evento {registro.get('NPJ')}:")
                logar(str(e))
                logar(traceback.format_exc())

                continue

            finally:
                relatorio_dados.append(registro)

    except Exception as e:
        logar("Erro geral dentro da funcao principal.")
        logar(str(e))
        logar(traceback.format_exc())

    return dados_processos, relatorio_dados


# ==========================================================
# EXECUCAO
# ==========================================================

def main():
    driver = None
    caminho_relatorio = None

    try:
        logar("Iniciando execucao do robo de baixa de distribuicoes.")

        driver, wait = iniciar_navegador()

        logar("Navegador iniciado com sucesso.")

        login_manual(driver)

        logar("Login manual realizado.")

        user = get_logged_user(driver)

        matricula = user.get("chave")
        nome = user.get("nome")

        if not matricula:
            raise Exception("Nao foi possivel obter a matricula do usuario logado.")

        logar(f"Usuario logado: {nome}")
        logar(f"Matricula: {matricula}")

        dados_processos, relatorio_dados = organizar_tramitacao_agenda_gerente(driver, matricula)

        caminho_relatorio = salvar_relatorio_excel(relatorio_dados)

        mensagem = (
            "Execucao finalizada.\n\n"
            f"Relatorio: {caminho_relatorio if caminho_relatorio else 'Nao gerado'}\n\n"
            f"Log: {CAMINHO_LOG}"
        )

        logar("Execucao finalizada.")

        mostrar_mensagem_final(
            titulo="Execucao finalizada",
            mensagem=mensagem,
            tipo="info"
        )

    except Exception as e:
        logar("Erro geral na execucao.")
        logar(str(e))
        logar(traceback.format_exc())

        mostrar_mensagem_final(
            titulo="Erro geral na execucao",
            mensagem=(
                "Ocorreu um erro geral na execucao.\n\n"
                f"Erro: {str(e)}\n\n"
                f"Verifique o log em:\n{CAMINHO_LOG}"
            ),
            tipo="erro"
        )

    finally:
        if driver:
            try:
                driver.quit()
                logar("Navegador fechado.")
            except Exception:
                pass


if __name__ == "__main__":
    main()