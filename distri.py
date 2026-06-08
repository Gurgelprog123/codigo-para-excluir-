from DijurLib.api.base import post_api_navegador, get_api_navegador, put_api_navegador

# Biblioteca Dijur
from DijurLib.portal.login import login_manual
from DijurLib.api.helpers import get_logged_user

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.edge.service import Service

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

from datetime import datetime
from tkinter import messagebox
from tkinter.scrolledtext import ScrolledText

import tkinter as tk
import pandas as pd
import os
import time
import re
import base64
import unicodedata
import json
import traceback


# ==========================================================
# CONFIGURACOES GERAIS
# ==========================================================

TEMPO = 1.5

# Mantive como caracteres porque era assim que seu codigo original validava.
# Se quiser validar por linhas reais, precisa trocar len(texto) por len(texto.splitlines()).
LIMITE_CARACTERES_PUBLICACAO = 3000

TEXTO_FIXO_DISTRIBUICAO_ADVOGADO = "NPJ distribuído para o advogado nesta data."

URL_CONSULTAR_EVENTOS = "https://juridico.intranet.bb.com.br/paj/resources/app/v1/agenda/consultar/evento/compromisso/responsavel/"
URL_ALTERAR_ESTADO_EVENTO = "https://juridico.intranet.bb.com.br/paj/resources/app/v1/agenda/evento/estado"
URL_REGISTRAR_OBSERVACAO = "https://juridico.intranet.bb.com.br/paj/resources/app/v1/agenda/evento/observacao"
URL_PUBLICACAO = "https://juridico.intranet.bb.com.br/paj/resources/app/v1/agenda/publicacao"


# ==========================================================
# CLASSES DE CONTROLE
# ==========================================================

class ExecucaoCancelada(Exception):
    pass


# ==========================================================
# NAVEGADOR
# ==========================================================

def iniciar_navegador(headless: bool = False, wait_time: int = 30, suppress_browser_ui: bool = False):
    """
    Inicializa o navegador Edge com as opcoes especificadas.
    """

    options = webdriver.EdgeOptions()

    if headless:
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    prefs = {
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True
    }

    if suppress_browser_ui:
        options.add_argument("--disable-notifications")
        options.add_argument("--no-download-notification")

        prefs.update({
            "browser.show_hub_popup_on_download_start": False,
            "user_experience_metrics.personalization_data_consent_enabled": True
        })

    options.add_experimental_option("prefs", prefs)

    try:
        if os.path.exists("C:\\EdgeDriver\\msedgedriver.exe"):
            service = Service(executable_path="C:\\EdgeDriver\\msedgedriver.exe")
            driver = webdriver.Edge(service=service, options=options)
        else:
            driver = webdriver.Edge(options=options)

    except Exception as e:
        msg = str(e)

        match = re.search(
            r"only supports Microsoft Edge version (\d+).*?Current browser version is (\d+\.\d+\.\d+\.\d+)",
            msg,
            re.DOTALL
        )

        with open("navegador_erro_log.txt", "w", encoding="utf-8") as f:
            if match:
                versao_suportada = match.group(1)
                versao_atual = match.group(2)

                f.write(
                    "Incompatibilidade de versao do Edge WebDriver.\n"
                    f"Driver suporta apenas a versao: {versao_suportada}\n"
                    f"Sua versao atual do Edge e: {versao_atual}\n"
                    "Atualize o msedgedriver.exe conforme a versao atual do navegador.\n"
                )

                try:
                    pdf_b64 = b"""
                    JJABDUIWDVBAPWVWI
                    """
                    nome_arquivo = "Manual Baixar Driver.pdf"

                    with open(nome_arquivo, "wb") as arquivo_pdf:
                        arquivo_pdf.write(base64.b64decode(pdf_b64))

                except Exception:
                    f.write("\nNao foi possivel gerar o PDF de orientacao automaticamente.\n")

            else:
                f.write(f"Ocorreu um erro ao abrir o navegador:\n{msg}")

        raise Exception("Erro de navegador")

    driver.maximize_window()

    wait = WebDriverWait(driver, wait_time)

    return driver, wait


# ==========================================================
# TKINTER
# ==========================================================

root = tk.Tk()
root.withdraw()

janela_temp = tk.Toplevel(root)
janela_temp.withdraw()
janela_temp.attributes("-topmost", True)


def serializar_objeto(objeto, limite=5000):
    """
    Transforma resposta/payload em texto legivel para janela e relatorio.
    """

    try:
        texto = json.dumps(objeto, ensure_ascii=False, indent=2, default=str)
    except Exception:
        texto = str(objeto)

    if texto is None:
        texto = ""

    if len(texto) > limite:
        texto = texto[:limite] + "\n\n... TEXTO CORTADO PARA VISUALIZACAO ..."

    return texto


def centralizar_janela(janela, largura=900, altura=650):
    """
    Centraliza uma janela tkinter.
    """

    janela.update_idletasks()

    largura_tela = janela.winfo_screenwidth()
    altura_tela = janela.winfo_screenheight()

    x = int((largura_tela / 2) - (largura / 2))
    y = int((altura_tela / 2) - (altura / 2))

    janela.geometry(f"{largura}x{altura}+{x}+{y}")


def mostrar_janela_requisicao(
    titulo,
    etapa,
    metodo,
    url,
    payload=None,
    resposta=None,
    texto_extra=None
):
    """
    Mostra janela depois de cada requisicao.
    O codigo so continua quando clicar em Continuar.
    """

    janela = tk.Toplevel(root)
    janela.title(titulo)
    janela.attributes("-topmost", True)
    janela.grab_set()

    centralizar_janela(janela, largura=950, altura=700)

    frame_principal = tk.Frame(janela)
    frame_principal.pack(fill="both", expand=True, padx=10, pady=10)

    label_titulo = tk.Label(
        frame_principal,
        text=etapa,
        font=("Arial", 13, "bold"),
        anchor="w",
        justify="left"
    )
    label_titulo.pack(fill="x", pady=(0, 10))

    texto = ScrolledText(frame_principal, wrap="word", font=("Consolas", 10))
    texto.pack(fill="both", expand=True)

    conteudo = []
    conteudo.append(f"ETAPA:\n{etapa}\n")
    conteudo.append(f"METODO:\n{metodo}\n")
    conteudo.append(f"URL:\n{url}\n")

    if payload is not None:
        conteudo.append("PAYLOAD ENVIADO:\n")
        conteudo.append(serializar_objeto(payload, limite=3000))
        conteudo.append("\n")

    if resposta is not None:
        conteudo.append("RESPOSTA RECEBIDA:\n")
        conteudo.append(serializar_objeto(resposta, limite=5000))
        conteudo.append("\n")

    if texto_extra:
        conteudo.append("OBSERVACAO:\n")
        conteudo.append(str(texto_extra))
        conteudo.append("\n")

    texto.insert("1.0", "\n".join(conteudo))
    texto.config(state="disabled")

    decisao = {"continuar": False}

    def continuar():
        decisao["continuar"] = True
        janela.destroy()

    def parar():
        decisao["continuar"] = False
        janela.destroy()

    frame_botoes = tk.Frame(frame_principal)
    frame_botoes.pack(fill="x", pady=(10, 0))

    btn_continuar = tk.Button(
        frame_botoes,
        text="Continuar para a proxima etapa",
        command=continuar,
        width=35,
        height=2
    )
    btn_continuar.pack(side="left", padx=(0, 10))

    btn_parar = tk.Button(
        frame_botoes,
        text="Parar execucao",
        command=parar,
        width=20,
        height=2
    )
    btn_parar.pack(side="left")

    janela.protocol("WM_DELETE_WINDOW", parar)

    janela.wait_window()

    if not decisao["continuar"]:
        raise ExecucaoCancelada("Execucao cancelada pelo usuario na janela de requisicao.")


def confirmar_baixa_manual(npj, cod_evento, regra, motivo, detalhes):
    """
    Janela exibida antes do PUT.
    O PUT e a baixa real, entao essa confirmacao vem antes da requisicao.
    """

    janela = tk.Toplevel(root)
    janela.title("CONFIRMAR BAIXA DO EVENTO")
    janela.attributes("-topmost", True)
    janela.grab_set()

    centralizar_janela(janela, largura=850, altura=550)

    frame_principal = tk.Frame(janela)
    frame_principal.pack(fill="both", expand=True, padx=10, pady=10)

    label = tk.Label(
        frame_principal,
        text="ATENCAO: a proxima requisicao e um PUT e vai concluir o evento.",
        font=("Arial", 13, "bold"),
        fg="red",
        anchor="w",
        justify="left"
    )
    label.pack(fill="x", pady=(0, 10))

    texto = ScrolledText(frame_principal, wrap="word", font=("Consolas", 10))
    texto.pack(fill="both", expand=True)

    conteudo = (
        f"NPJ:\n{npj}\n\n"
        f"Codigo do evento:\n{cod_evento}\n\n"
        f"Regra identificada:\n{regra}\n\n"
        f"Motivo da conclusao:\n{motivo}\n\n"
        f"Detalhes:\n{detalhes}\n\n"
        "Escolha uma opcao:\n"
        "- Confirmar baixa: realiza o PUT e conclui o evento.\n"
        "- Pular este evento: nao baixa este evento e vai para o proximo.\n"
        "- Parar execucao: encerra o processamento.\n"
    )

    texto.insert("1.0", conteudo)
    texto.config(state="disabled")

    decisao = {"acao": "parar"}

    def confirmar():
        decisao["acao"] = "confirmar"
        janela.destroy()

    def pular():
        decisao["acao"] = "pular"
        janela.destroy()

    def parar():
        decisao["acao"] = "parar"
        janela.destroy()

    frame_botoes = tk.Frame(frame_principal)
    frame_botoes.pack(fill="x", pady=(10, 0))

    btn_confirmar = tk.Button(
        frame_botoes,
        text="Confirmar baixa",
        command=confirmar,
        width=20,
        height=2
    )
    btn_confirmar.pack(side="left", padx=(0, 10))

    btn_pular = tk.Button(
        frame_botoes,
        text="Pular este evento",
        command=pular,
        width=20,
        height=2
    )
    btn_pular.pack(side="left", padx=(0, 10))

    btn_parar = tk.Button(
        frame_botoes,
        text="Parar execucao",
        command=parar,
        width=20,
        height=2
    )
    btn_parar.pack(side="left")

    janela.protocol("WM_DELETE_WINDOW", parar)

    janela.wait_window()

    return decisao["acao"]


# ==========================================================
# FUNCOES AUXILIARES
# ==========================================================

def normalizar_texto(texto):
    """
    Normaliza texto para comparacoes mais seguras.
    """

    if texto is None:
        return ""

    texto = str(texto).lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(char for char in texto if unicodedata.category(char) != "Mn")
    texto = re.sub(r"\s+", " ", texto).strip()

    return texto


def eh_texto_fixo_distribuicao_advogado(texto_descricao_compromisso):
    """
    Verifica a primeira forma de baixa:
    texto fixo de NPJ distribuido para advogado.
    """

    texto_recebido = normalizar_texto(texto_descricao_compromisso)
    texto_esperado = normalizar_texto(TEXTO_FIXO_DISTRIBUICAO_ADVOGADO)

    return texto_recebido == texto_esperado


def extrair_turma_distribuicao_tst(conteudo_publicacao):
    """
    Verifica a segunda forma de baixa:
    publicacao curta do TST informando distribuicao para turma.
    """

    texto = normalizar_texto(conteudo_publicacao)

    padrao = (
        r"\bdistribuid[oa]\s+para\s+"
        r"(?:a\s+)?"
        r"(\d{1,2})"
        r"\s*(?:a|o|ª|º)?\s*"
        r"turma\b"
    )

    match = re.search(padrao, texto, flags=re.IGNORECASE)

    if not match:
        return None

    numero_turma = int(match.group(1))

    if numero_turma <= 0:
        return None

    return f"{numero_turma:02d} TURMA"


def resposta_sucesso(resposta):
    """
    Padroniza a validacao de sucesso das respostas da API.
    """

    if not resposta:
        return False

    if resposta.get("type") == "SUCCESS":
        return True

    if resposta.get("status") == "OK":
        return True

    return False


def resumo_publicacao(texto, limite=500):
    """
    Gera resumo curto da publicacao para o relatorio.
    """

    if not texto:
        return ""

    texto = re.sub(r"\s+", " ", str(texto)).strip()

    if len(texto) > limite:
        return texto[:limite] + "..."

    return texto


def criar_registro_relatorio(evento, indice_evento):
    """
    Cria registro base para auditoria do evento.
    """

    return {
        "Data Hora Execucao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "Indice Evento": indice_evento,
        "Codigo Evento": evento.get("codigoEventoJuridico"),
        "NPJ": evento.get("numeroProcessoFormatado"),
        "Texto Descricao Compromisso": evento.get("textoDescricaoCompromisso"),
        "Status Final": "NAO PROCESSADO",
        "Regra Identificada": "",
        "Acao Executada": "",
        "Motivo": "",
        "Tramitacao Identificada": "",
        "Numero Processo": "",
        "Numero Sequencial Publicacao": "",
        "Qtd Caracteres Publicacao": "",
        "Qtd Linhas Publicacao": "",
        "Resumo Publicacao": "",
        "Resposta PUT": "",
        "Resposta Observacao": "",
        "Erro": ""
    }


def executar_post_api(driver, url, payload, etapa):
    """
    Executa POST normal usando a biblioteca Dijur e mostra janela depois.
    """

    resposta = post_api_navegador(driver, url, payload)

    mostrar_janela_requisicao(
        titulo="Requisicao POST executada",
        etapa=etapa,
        metodo="POST",
        url=url,
        payload=payload,
        resposta=resposta
    )

    return resposta


def executar_get_api(driver, url, etapa):
    """
    Executa GET usando a biblioteca Dijur e mostra janela depois.
    """

    resposta = get_api_navegador(driver, url)

    mostrar_janela_requisicao(
        titulo="Requisicao GET executada",
        etapa=etapa,
        metodo="GET",
        url=url,
        payload=None,
        resposta=resposta
    )

    return resposta


def executar_put_api(driver, url, payload, etapa):
    """
    Executa PUT usando a biblioteca Dijur e mostra janela depois.
    """

    resposta = put_api_navegador(driver, url, payload)

    mostrar_janela_requisicao(
        titulo="Requisicao PUT executada",
        etapa=etapa,
        metodo="PUT",
        url=url,
        payload=payload,
        resposta=resposta,
        texto_extra="Essa requisicao e a baixa/conclusao do evento."
    )

    return resposta


def post_api_navegador_plain(driver, api_url: str, payload: str, etapa: str, max_attempts: int = 3):
    """
    Faz POST enviando payload como texto puro.
    Usado na consulta da publicacao.
    """

    def parse_service_response(texto: str) -> dict:
        padrao = r"status=(\w+).*?messages?=\[([^\]]*)\].*?data=(.*)"
        match = re.search(padrao, texto, flags=re.DOTALL)

        if match:
            status, message, data = match.groups()

            return {
                "status": status,
                "message": message,
                "data": data
            }

        return {"raw": texto}

    attempts = 0

    while attempts < max_attempts:
        try:
            print("Iniciando requisicao POST plain...")

            response_data = driver.execute_async_script(
                """
                const api_url = arguments[0];
                const payload = arguments[1];
                const callback = arguments[arguments.length - 1];

                fetch(api_url, {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: payload,
                    credentials: 'same-origin'
                })
                .then(response => {
                    if (!response.ok) {
                        throw new Error('Network response was not ok: ' + response.statusText);
                    }
                    return response.text();
                })
                .then(text => {
                    try {
                        const parsed = JSON.parse(text);
                        callback(parsed);
                    } catch (err) {
                        callback({ rawText: text });
                    }
                })
                .catch(error => callback({ error: error.toString() }));
                """,
                api_url,
                payload
            )

            if "error" in response_data:
                resposta_erro = {"error": response_data["error"]}

                mostrar_janela_requisicao(
                    titulo="Requisicao POST plain executada com erro",
                    etapa=f"{etapa} - tentativa {attempts + 1}",
                    metodo="POST PLAIN",
                    url=api_url,
                    payload=payload,
                    resposta=resposta_erro
                )

                return None

            if "rawText" in response_data:
                parsed = parse_service_response(response_data["rawText"])

                mostrar_janela_requisicao(
                    titulo="Requisicao POST plain executada",
                    etapa=f"{etapa} - tentativa {attempts + 1}",
                    metodo="POST PLAIN",
                    url=api_url,
                    payload=payload,
                    resposta=parsed
                )

                return parsed

            mostrar_janela_requisicao(
                titulo="Requisicao POST plain executada",
                etapa=f"{etapa} - tentativa {attempts + 1}",
                metodo="POST PLAIN",
                url=api_url,
                payload=payload,
                resposta=response_data
            )

            return response_data

        except ExecucaoCancelada:
            raise

        except Exception as e:
            resposta_erro = {
                "erro": str(e),
                "traceback": traceback.format_exc()
            }

            mostrar_janela_requisicao(
                titulo="Erro na requisicao POST plain",
                etapa=f"{etapa} - tentativa {attempts + 1}",
                metodo="POST PLAIN",
                url=api_url,
                payload=payload,
                resposta=resposta_erro
            )

        attempts += 1

        if attempts < max_attempts:
            print("Tentando novamente em 1 segundo...")
            time.sleep(1)

    print("Numero maximo de tentativas alcancado. Retornando None.")
    return None


# ==========================================================
# API / FLUXO DE NEGOCIO
# ==========================================================

def buscar_eventos_pendentes(driver, matricula):
    """
    Busca eventos pendentes da agenda.
    """

    todos_eventos = []
    posicao = 1

    while True:
        payload = {
            "posicao": posicao,
            "responsavel": matricula,
            "situacao": 5
        }

        resposta = executar_post_api(
            driver=driver,
            url=URL_CONSULTAR_EVENTOS,
            payload=payload,
            etapa=f"Buscar eventos pendentes da agenda - posicao {posicao}"
        )

        if resposta is None or resposta.get("status") != "OK":
            print("Status na busca de publicacoes veio None ou diferente de OK.")
            break

        lista_eventos = resposta.get("data", {}).get("listaEvento", [])

        if not lista_eventos:
            break

        todos_eventos.extend(lista_eventos)

        if len(lista_eventos) < 50:
            break

        posicao += 50

    return todos_eventos


def obter_numero_processo(driver, num_processo_formatado):
    """
    Obtem numeroProcesso a partir do NPJ formatado.
    """

    numero_processo_sem_barra = num_processo_formatado.replace("/", "")
    partes = numero_processo_sem_barra.split("-")

    if len(partes) != 2:
        print(f"NPJ em formato inesperado: {num_processo_formatado}")
        return None, None

    npj_base = partes[0]

    try:
        npj_sufixo = int(partes[1])

    except ValueError:
        print(f"Variacao do NPJ invalida: {num_processo_formatado}")
        return None, None

    url_numeroprocesso = (
        "https://juridico.intranet.bb.com.br/paj/resources/app/v1/"
        f"processo/consulta/{npj_base}/{npj_sufixo}/0"
    )

    resposta_numeroprocesso = executar_get_api(
        driver=driver,
        url=url_numeroprocesso,
        etapa=f"Buscar numeroProcesso do NPJ {num_processo_formatado}"
    )

    if resposta_numeroprocesso is None or resposta_numeroprocesso.get("status") != "OK":
        print("Status da chamada para pegar numeroProcesso diferente de OK.")
        return None, resposta_numeroprocesso

    lista_ocorrencia = resposta_numeroprocesso.get("data", {}).get("listaOcorrencia", [])

    if not lista_ocorrencia:
        print(f"Nenhuma ocorrencia encontrada para o processo: {num_processo_formatado}")
        return None, resposta_numeroprocesso

    numero_processo = lista_ocorrencia[0].get("numeroProcesso")

    if not numero_processo:
        print(f"numeroProcesso nao encontrado para: {num_processo_formatado}")
        return None, resposta_numeroprocesso

    return numero_processo, resposta_numeroprocesso


def obter_numero_sequencial_publicacao(driver, cod_evento, matricula, npj):
    """
    Obtem numeroSequencialPublicacaoJudicial pelo codigo do evento.
    """

    url_detalhes = (
        "https://juridico.intranet.bb.com.br/paj/resources/app/v1/"
        f"agenda/evento/{cod_evento}/{matricula}"
    )

    resposta_detalhes = executar_get_api(
        driver=driver,
        url=url_detalhes,
        etapa=f"Buscar detalhes do evento {cod_evento} / NPJ {npj}"
    )

    if resposta_detalhes is None or resposta_detalhes.get("status") != "OK":
        print("Status da chamada para pegar numeroSequencialPublicacaoJudicial diferente de OK.")
        return None, resposta_detalhes

    numero_sequencial = resposta_detalhes.get("data", {}).get("numeroSequencialPublicacaoJudicial")

    if not numero_sequencial:
        print(f"numeroSequencialPublicacaoJudicial nao encontrado para evento: {cod_evento}")
        return None, resposta_detalhes

    return numero_sequencial, resposta_detalhes


def obter_conteudo_publicacao(driver, cod_evento, numero_sequencial_publicacao, npj):
    """
    Busca o conteudo da publicacao.
    """

    payload_publicacao = f"{numero_sequencial_publicacao};{cod_evento}"

    resposta_publicacao = post_api_navegador_plain(
        driver=driver,
        api_url=URL_PUBLICACAO,
        payload=payload_publicacao,
        etapa=f"Buscar conteudo da publicacao do NPJ {npj}"
    )

    if resposta_publicacao is None:
        print("Resposta da chamada de publicacao veio None.")
        return None, resposta_publicacao

    status_publicacao = resposta_publicacao.get("status")
    type_publicacao = resposta_publicacao.get("type")

    if status_publicacao != "OK" and type_publicacao != "SUCCESS":
        print("Status da chamada para pegar conteudo diferente de OK/SUCCESS.")
        print(resposta_publicacao)
        return None, resposta_publicacao

    dados_publicacao = resposta_publicacao.get("data")

    if isinstance(dados_publicacao, dict):
        conteudo_publicacao = dados_publicacao.get("textoPublicacao")

    elif isinstance(dados_publicacao, str):
        conteudo_publicacao = dados_publicacao

    else:
        print("Resposta da publicacao veio em formato inesperado:")
        print(resposta_publicacao)
        return None, resposta_publicacao

    if not conteudo_publicacao:
        print("Publicacao sem texto.")
        return None, resposta_publicacao

    return conteudo_publicacao, resposta_publicacao


def concluir_evento(driver, matricula, cod_evento, motivo, npj, regra):
    """
    Realiza o PUT que conclui o evento.
    Esse e o ponto onde a baixa acontece.
    """

    put_payload = {
        "chave": matricula,
        "codigoEventoJuridico": cod_evento,
        "codigoTipoEstadoCompromisso": 3,
        "inclurAndamentoAutomatico": True,
        "justificativaConclusao": motivo,
    }

    resposta_put = executar_put_api(
        driver=driver,
        url=URL_ALTERAR_ESTADO_EVENTO,
        payload=put_payload,
        etapa=f"PUT de baixa do evento {cod_evento} / NPJ {npj} / Regra {regra}"
    )

    return resposta_put


def registrar_observacao_conclusao(driver, cod_evento, texto_observacao, npj):
    """
    Registra observacao depois que a baixa foi feita com sucesso.
    """

    payload_post_descricao = {
        "codigoEventoJuridico": cod_evento,
        "textoObservacaoCompromisso": texto_observacao
    }

    resposta_observacao = executar_post_api(
        driver=driver,
        url=URL_REGISTRAR_OBSERVACAO,
        payload=payload_post_descricao,
        etapa=f"Registrar observacao de conclusao do evento {cod_evento} / NPJ {npj}"
    )

    return resposta_observacao


# ==========================================================
# RELATORIO
# ==========================================================

def salvar_relatorio_excel(relatorio_dados):
    """
    Gera relatorio Excel com auditoria completa da execucao.
    """

    if not relatorio_dados:
        print("\nNenhum dado coletado para gerar o relatorio.")
        return

    df_relatorio = pd.DataFrame(relatorio_dados)

    data_hoje = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    nome_arquivo = f"Relatorio auditoria baixas distribuicao {data_hoje}.xlsx"
    caminho_arquivo = os.path.join(os.getcwd(), nome_arquivo)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Auditoria Baixas"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, column_name in enumerate(df_relatorio.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment

        for row_idx, row in enumerate(dataframe_to_rows(df_relatorio, index=False, header=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = alignment

                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    tamanho = len(str(cell.value))

                    if tamanho > max_length:
                        max_length = tamanho

                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 35

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(caminho_arquivo)

        print(f"\nRelatorio de auditoria salvo com sucesso em: {caminho_arquivo}")

        messagebox.showinfo(
            "Relatorio gerado",
            f"Relatorio de auditoria salvo com sucesso em:\n\n{caminho_arquivo}",
            parent=janela_temp
        )

    except Exception as e:
        print("\nErro ao salvar o relatorio Excel:")
        print(e)

        messagebox.showerror(
            "Erro ao salvar relatorio",
            f"Erro ao salvar o relatorio Excel:\n\n{str(e)}",
            parent=janela_temp
        )


# ==========================================================
# FUNCAO PRINCIPAL
# ==========================================================

def organizar_tramitacao_agenda_gerente(driver, matricula):
    """
    Organiza a agenda e realiza baixa de publicacoes pendentes de prazo
    em duas situacoes:

    1. Evento com texto fixo:
       NPJ distribuido para o advogado nesta data.

    2. Publicacao curta do TST:
       texto com ate 3000 caracteres e regex de distribuicao para turma.
    """

    dados_processos = []
    relatorio_dados = []
    eventos_baixados = set()

    try:
        todos_eventos = buscar_eventos_pendentes(driver, matricula)

        if not todos_eventos:
            messagebox.showinfo(
                "Atencao",
                "Nenhum evento encontrado ou erro na resposta.",
                parent=janela_temp
            )

            time.sleep(5)
            return dados_processos

        print("Agenda listada.")
        print(f"Quantidade de eventos encontrados: {len(todos_eventos)}")

        for indice_evento, evento in enumerate(todos_eventos, 1):
            registro = criar_registro_relatorio(evento, indice_evento)

            try:
                cod_evento = evento.get("codigoEventoJuridico")
                num_processo_formatado = evento.get("numeroProcessoFormatado")
                texto_descricao_compromisso = evento.get("textoDescricaoCompromisso")

                if not cod_evento:
                    registro["Status Final"] = "NAO BAIXADO - SEM CODIGO EVENTO"
                    registro["Acao Executada"] = "Nenhuma"
                    registro["Motivo"] = "Evento sem codigoEventoJuridico."
                    continue

                if cod_evento in eventos_baixados:
                    registro["Status Final"] = "NAO BAIXADO - EVENTO JA BAIXADO NA EXECUCAO"
                    registro["Acao Executada"] = "Nenhuma"
                    registro["Motivo"] = "Evento ja foi baixado anteriormente nesta mesma execucao."
                    continue

                # ==================================================
                # REGRA 1: BAIXA PELO TEXTO FIXO DO COMPROMISSO
                # ==================================================

                if eh_texto_fixo_distribuicao_advogado(texto_descricao_compromisso):
                    registro["Regra Identificada"] = "Distribuicao interna - texto fixo"
                    registro["Tramitacao Identificada"] = "DISTRIBUICAO PARA ADVOGADO"

                    print("Evento identificado pela regra 1: texto fixo de distribuicao para advogado.")

                    decisao = confirmar_baixa_manual(
                        npj=num_processo_formatado,
                        cod_evento=cod_evento,
                        regra="Distribuicao interna - texto fixo",
                        motivo="Distribuição.",
                        detalhes=(
                            "O campo textoDescricaoCompromisso veio exatamente como:\n"
                            f"{TEXTO_FIXO_DISTRIBUICAO_ADVOGADO}\n\n"
                            "Essa e a primeira forma de distribuicao informada."
                        )
                    )

                    if decisao == "pular":
                        registro["Status Final"] = "PULADO MANUALMENTE"
                        registro["Acao Executada"] = "Nenhuma"
                        registro["Motivo"] = "Usuario escolheu pular este evento antes do PUT."
                        continue

                    if decisao == "parar":
                        registro["Status Final"] = "EXECUCAO CANCELADA"
                        registro["Acao Executada"] = "Nenhuma"
                        registro["Motivo"] = "Usuario escolheu parar a execucao antes do PUT."
                        raise ExecucaoCancelada("Execucao cancelada antes do PUT da regra 1.")

                    resposta_put = concluir_evento(
                        driver=driver,
                        matricula=matricula,
                        cod_evento=cod_evento,
                        motivo="Distribuição.",
                        npj=num_processo_formatado,
                        regra="Distribuicao interna - texto fixo"
                    )

                    registro["Resposta PUT"] = serializar_objeto(resposta_put, limite=1500)

                    if not resposta_sucesso(resposta_put):
                        registro["Status Final"] = "ERRO - FALHA NO PUT"
                        registro["Acao Executada"] = "Tentou realizar baixa"
                        registro["Motivo"] = "PUT retornou erro ou resposta diferente de sucesso."
                        print("Falha ao concluir evento pela regra 1.")
                        break

                    eventos_baixados.add(cod_evento)

                    time.sleep(TEMPO)

                    resposta_observacao = registrar_observacao_conclusao(
                        driver=driver,
                        cod_evento=cod_evento,
                        npj=num_processo_formatado,
                        texto_observacao=(
                            "Evento concluido. Motivo: Distribuição. "
                            "Regra: texto fixo NPJ distribuído para o advogado."
                        )
                    )

                    registro["Resposta Observacao"] = serializar_objeto(resposta_observacao, limite=1500)

                    if not resposta_sucesso(resposta_observacao):
                        registro["Status Final"] = "ERRO - FALHA AO REGISTRAR OBSERVACAO"
                        registro["Acao Executada"] = "Baixou evento, mas falhou ao registrar observacao"
                        registro["Motivo"] = "A baixa foi feita, mas a observacao retornou erro."
                        print("Falha ao registrar observacao pela regra 1.")
                        break

                    registro["Status Final"] = "BAIXADO - DISTRIBUICAO INTERNA"
                    registro["Acao Executada"] = "Baixa realizada por PUT e observacao registrada"
                    registro["Motivo"] = "Texto fixo de distribuicao interna identificado."

                    time.sleep(TEMPO)

                    continue

                # ==================================================
                # REGRA 2: BAIXA POR PUBLICACAO CURTA DO TST
                # ==================================================

                if not num_processo_formatado:
                    registro["Status Final"] = "NAO BAIXADO - SEM NPJ"
                    registro["Acao Executada"] = "Nenhuma"
                    registro["Motivo"] = "Evento sem numeroProcessoFormatado."
                    continue

                numero_processo, resposta_numero_processo = obter_numero_processo(
                    driver=driver,
                    num_processo_formatado=num_processo_formatado
                )

                if not numero_processo:
                    registro["Status Final"] = "NAO BAIXADO - PROCESSO NAO LOCALIZADO"
                    registro["Acao Executada"] = "Consultou processo"
                    registro["Motivo"] = "Nao foi possivel obter numeroProcesso."
                    continue

                registro["Numero Processo"] = numero_processo

                time.sleep(TEMPO)

                numero_sequencial_publicacao, resposta_detalhes = obter_numero_sequencial_publicacao(
                    driver=driver,
                    cod_evento=cod_evento,
                    matricula=matricula,
                    npj=num_processo_formatado
                )

                if not numero_sequencial_publicacao:
                    registro["Status Final"] = "NAO BAIXADO - SEM PUBLICACAO JUDICIAL"
                    registro["Acao Executada"] = "Consultou detalhes do evento"
                    registro["Motivo"] = "Nao encontrou numeroSequencialPublicacaoJudicial."
                    continue

                registro["Numero Sequencial Publicacao"] = numero_sequencial_publicacao

                time.sleep(TEMPO)

                conteudo_publicacao, resposta_publicacao = obter_conteudo_publicacao(
                    driver=driver,
                    cod_evento=cod_evento,
                    numero_sequencial_publicacao=numero_sequencial_publicacao,
                    npj=num_processo_formatado
                )

                if not conteudo_publicacao:
                    registro["Status Final"] = "NAO BAIXADO - SEM TEXTO DA PUBLICACAO"
                    registro["Acao Executada"] = "Consultou publicacao"
                    registro["Motivo"] = "Nao foi possivel obter texto da publicacao."
                    continue

                quantidade_caracteres = len(conteudo_publicacao)
                quantidade_linhas = len(str(conteudo_publicacao).splitlines())

                registro["Qtd Caracteres Publicacao"] = quantidade_caracteres
                registro["Qtd Linhas Publicacao"] = quantidade_linhas
                registro["Resumo Publicacao"] = resumo_publicacao(conteudo_publicacao, limite=500)

                if quantidade_caracteres > LIMITE_CARACTERES_PUBLICACAO:
                    registro["Status Final"] = "NAO BAIXADO - PUBLICACAO EXTENSA"
                    registro["Acao Executada"] = "Consultou publicacao, mas nao baixou"
                    registro["Motivo"] = (
                        f"Publicacao com {quantidade_caracteres} caracteres. "
                        f"Limite configurado: {LIMITE_CARACTERES_PUBLICACAO}."
                    )

                    print(
                        "Publicacao ignorada por tamanho. "
                        f"Caracteres: {quantidade_caracteres}. "
                        f"Limite: {LIMITE_CARACTERES_PUBLICACAO}."
                    )

                    continue

                tramitacao_distribuida = extrair_turma_distribuicao_tst(conteudo_publicacao)

                if not tramitacao_distribuida:
                    registro["Status Final"] = "NAO BAIXADO - NAO ERA DISTRIBUICAO"
                    registro["Regra Identificada"] = "Nenhuma"
                    registro["Acao Executada"] = "Consultou publicacao, mas nao baixou"
                    registro["Motivo"] = "Publicacao curta, mas sem regex de distribuicao para turma."
                    print("Publicacao curta, mas sem regex de distribuicao para turma.")
                    continue

                registro["Regra Identificada"] = "Distribuicao TST - publicacao curta"
                registro["Tramitacao Identificada"] = tramitacao_distribuida

                print(f"Evento identificado pela regra 2: publicacao TST distribuida para {tramitacao_distribuida}.")

                decisao = confirmar_baixa_manual(
                    npj=num_processo_formatado,
                    cod_evento=cod_evento,
                    regra="Distribuicao TST - publicacao curta",
                    motivo="Distribuição.",
                    detalhes=(
                        f"A publicacao possui {quantidade_caracteres} caracteres.\n"
                        f"Limite configurado: {LIMITE_CARACTERES_PUBLICACAO} caracteres.\n\n"
                        f"Tramitacao identificada pela regex: {tramitacao_distribuida}\n\n"
                        "Resumo da publicacao:\n"
                        f"{resumo_publicacao(conteudo_publicacao, limite=1200)}"
                    )
                )

                if decisao == "pular":
                    registro["Status Final"] = "PULADO MANUALMENTE"
                    registro["Acao Executada"] = "Nenhuma"
                    registro["Motivo"] = "Usuario escolheu pular este evento antes do PUT."
                    continue

                if decisao == "parar":
                    registro["Status Final"] = "EXECUCAO CANCELADA"
                    registro["Acao Executada"] = "Nenhuma"
                    registro["Motivo"] = "Usuario escolheu parar a execucao antes do PUT."
                    raise ExecucaoCancelada("Execucao cancelada antes do PUT da regra 2.")

                resposta_put = concluir_evento(
                    driver=driver,
                    matricula=matricula,
                    cod_evento=cod_evento,
                    motivo="Distribuição.",
                    npj=num_processo_formatado,
                    regra="Distribuicao TST - publicacao curta"
                )

                registro["Resposta PUT"] = serializar_objeto(resposta_put, limite=1500)

                if not resposta_sucesso(resposta_put):
                    registro["Status Final"] = "ERRO - FALHA NO PUT"
                    registro["Acao Executada"] = "Tentou realizar baixa"
                    registro["Motivo"] = "PUT retornou erro ou resposta diferente de sucesso."
                    print("Falha ao concluir evento pela regra 2.")
                    break

                eventos_baixados.add(cod_evento)

                time.sleep(TEMPO)

                resposta_observacao = registrar_observacao_conclusao(
                    driver=driver,
                    cod_evento=cod_evento,
                    npj=num_processo_formatado,
                    texto_observacao=(
                        "Evento concluido. Motivo: Distribuição. "
                        f"Regra: publicação TST curta. Tramitação identificada: {tramitacao_distribuida}."
                    )
                )

                registro["Resposta Observacao"] = serializar_objeto(resposta_observacao, limite=1500)

                if not resposta_sucesso(resposta_observacao):
                    registro["Status Final"] = "ERRO - FALHA AO REGISTRAR OBSERVACAO"
                    registro["Acao Executada"] = "Baixou evento, mas falhou ao registrar observacao"
                    registro["Motivo"] = "A baixa foi feita, mas a observacao retornou erro."
                    print("Falha ao registrar observacao pela regra 2.")
                    break

                registro["Status Final"] = "BAIXADO - DISTRIBUICAO TST"
                registro["Acao Executada"] = "Baixa realizada por PUT e observacao registrada"
                registro["Motivo"] = "Publicacao curta do TST com regex de distribuicao para turma."

                time.sleep(TEMPO)

            except ExecucaoCancelada:
                registro["Status Final"] = "EXECUCAO CANCELADA"
                registro["Acao Executada"] = "Execucao interrompida pelo usuario"
                registro["Erro"] = "Execucao cancelada pelo usuario."
                raise

            except Exception as e:
                registro["Status Final"] = "ERRO"
                registro["Acao Executada"] = "Erro durante processamento do evento"
                registro["Erro"] = str(e)
                registro["Motivo"] = "Ocorreu erro inesperado no processamento deste evento."

                print(f"Erro ao processar evento {registro.get('NPJ')}:")
                print(e)
                print(traceback.format_exc())

                continue

            finally:
                relatorio_dados.append(registro)

    except ExecucaoCancelada as e:
        print(str(e))

        messagebox.showwarning(
            "Execucao cancelada",
            f"A execucao foi cancelada pelo usuario.\n\n{str(e)}",
            parent=janela_temp
        )

    finally:
        salvar_relatorio_excel(relatorio_dados)

    return dados_processos


# ==========================================================
# EXECUCAO
# ==========================================================

driver = None

try:
    driver, wait = iniciar_navegador()

    login_manual(driver)

    user = get_logged_user(driver)

    matricula = user.get("chave")
    nome = user.get("nome")

    if not matricula:
        raise Exception("Nao foi possivel obter a matricula do usuario logado.")

    print(f"Usuario logado: {nome}")
    print(f"Matricula: {matricula}")

    organizar_tramitacao_agenda_gerente(driver, matricula)

except ExecucaoCancelada as e:
    print(f"Execucao cancelada: {str(e)}")

except Exception as e:
    print("Erro geral na execucao:")
    print(e)
    print(traceback.format_exc())

    messagebox.showerror(
        "Erro geral na execucao",
        f"Ocorreu um erro geral:\n\n{str(e)}",
        parent=janela_temp
    )

finally:
    if driver:
        try:
            driver.quit()
        except Exception:
            pass